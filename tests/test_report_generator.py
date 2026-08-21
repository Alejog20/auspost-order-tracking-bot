import json
from datetime import date, datetime, timedelta

import pytest

import report_generator as rg
from models import TrackingItem


@pytest.fixture
def config():
    return rg.load_template_config()


@pytest.fixture
def sample_items():
    return [
        TrackingItem(
            tracking_number="ABC123",
            status="In transit",
            category="in_transit",
            last_scan_time=datetime.now(),
            expected_delivery=date.today() + timedelta(days=3),
            expected_delivery_label="Fri 7 - Mon 10 Aug",
        ),
        TrackingItem(
            tracking_number="XYZ789",
            status="Awaiting collection",
            category="awaiting_collection",
            last_scan_time=datetime.now(),
            collection_location="Mount Beauty LPO",
            collection_deadline=date.today() + timedelta(days=5),
        ),
    ]


def test_build_prompt_includes_tracking_numbers(config, sample_items):
    _, shipment_block = rg.build_prompt(sample_items, config)
    assert "ABC123" in shipment_block
    assert "XYZ789" in shipment_block


def test_build_prompt_awaiting_collection_includes_deadline_not_delivery_window(
    config,
    sample_items,
):
    _, shipment_block = rg.build_prompt(sample_items, config)
    collection_line = [l for l in shipment_block.split("\n") if "XYZ789" in l][0]

    assert "Mount Beauty LPO" in collection_line
    assert "expected delivery window" not in collection_line


def test_prepare_render_context_buckets_items_correctly(config, sample_items):
    narrative = {
        "summary_headline": "test_headline",
        "items": [
            {
                "tracking_number": "ABC123",
                "narrative": "in transit narrative",
            },
            {
                "tracking_number": "XYZ789",
                "narrative": "collection narrative",
            },
        ],
    }

    context = rg._prepare_render_context(sample_items, config, narrative)

    counts = {s["label"]: s["count"] for s in context["stats"]}

    assert counts["In transit"] == 1
    assert counts["Awaiting collection"] == 1
    assert counts["Delivered"] == 0


def test_render_report_includes_narrative_text(config, sample_items):
    narrative = {
        "summary_headline": "test headline",
        "items": [
            {
                "tracking_number": "ABC123",
                "narrative": "in transit narrative",
            },
            {
                "tracking_number": "XYZ789",
                "narrative": "collection narrative",
            },
        ],
    }

    html = rg._render_report(sample_items, config, narrative)

    assert "in transit narrative" in html
    assert "ABC123" in html


def test_log_report_writes_category(config, sample_items, tmp_path):
    log_path = tmp_path / "history.jsonl"

    rg.log_report(sample_items, config, log_path=log_path)

    entry = json.loads(log_path.read_text().splitlines()[0])

    categories = {
        i["tracking_number"]: i["category"]
        for i in entry["items"]
    }

    assert categories["ABC123"] == "in_transit"
    assert categories["XYZ789"] == "awaiting_collection"


def test_generate_report_returns_none_when_all_items_filtered(mocker, config):
    """
    If every item is a delivered shipment past drop_after_days, the drop-off
    filter empties the list before the Claude call. Must skip the call
    entirely (Claude rejects an empty user message) and signal "nothing to
    send" to the caller instead of raising.
    """
    stale_delivered = TrackingItem(
        tracking_number="OLD123",
        status="Delivered",
        category="delivered",
        last_scan_time=datetime.now() - timedelta(days=10),
    )
    mocker.patch(
        "history.filter_dropped_items",
        return_value=[],
    )
    mock_call_claude = mocker.patch("report_generator.call_claude")

    result = rg.generate_report([stale_delivered], template_path=None)

    assert result is None
    mock_call_claude.assert_not_called()


def test_generate_status_spreadsheet_returns_none_when_all_items_filtered(mocker, config, sample_items):
    mocker.patch("history.filter_dropped_items", return_value=[])

    result = rg.generate_status_spreadsheet(sample_items)

    assert result is None


def test_generate_status_spreadsheet_contains_tracking_rows(config, sample_items):
    from io import BytesIO

    from openpyxl import load_workbook

    xlsx_bytes = rg.generate_status_spreadsheet(sample_items)
    workbook = load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("Tracking number", "Status", "Category", "Needs attention", "Days since scan")

    by_tracking_number = {row[0]: row for row in rows[1:]}
    assert set(by_tracking_number) == {"ABC123", "XYZ789"}

    abc_row = by_tracking_number["ABC123"]
    assert abc_row[1] == "In transit"
    assert abc_row[2] == "in_transit"
    assert abc_row[3] == "No"
    assert abc_row[4] == 0


def test_generate_status_spreadsheet_excludes_items_the_drop_off_filter_removes(mocker, config, sample_items):
    from io import BytesIO

    from openpyxl import load_workbook

    mocker.patch("history.filter_dropped_items", return_value=sample_items[:1])

    xlsx_bytes = rg.generate_status_spreadsheet(sample_items)
    workbook = load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active

    tracking_numbers = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert tracking_numbers == {"ABC123"}


def test_call_claude_parses_json_response(mocker):
    """
    Mocked; doesn't hit the real API.
    This pattern can be reused for the Shopify and AusPost connectors.
    """

    fake_response = mocker.Mock()
    fake_response.content = [
        mocker.Mock(
            text='{"summary_headline":"ok","items":[]}'
        )
    ]

    mock_client = mocker.Mock()
    mock_client.messages.create.return_value = fake_response

    mocker.patch("anthropic.Anthropic", return_value=mock_client)

    result = rg.call_claude("system prompt", "shipment block")

    assert result == {
        "summary_headline": "ok",
        "items": [],
    }

    mock_client.messages.create.assert_called_once()


def test_call_claude_strips_markdown_code_fence(mocker):
    """
    Real Claude responses sometimes wrap JSON in a ```json ... ``` fence
    even when asked to respond as JSON only -- confirmed against a live call.
    """

    fake_response = mocker.Mock()
    fake_response.content = [
        mocker.Mock(
            text='```json\n{"summary_headline":"ok","items":[]}\n```'
        )
    ]

    mock_client = mocker.Mock()
    mock_client.messages.create.return_value = fake_response

    mocker.patch("anthropic.Anthropic", return_value=mock_client)

    result = rg.call_claude("system prompt", "shipment block")

    assert result == {
        "summary_headline": "ok",
        "items": [],
    }